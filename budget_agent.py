#!/usr/bin/env python3
"""
Project an FY2026-27 budget for the Marine Exchange of the SF Bay Region.

Inputs:
  - FY2024-25 P&L  (older year, "Budget vs. Actuals" QuickBooks export)
  - FY2025-26 P&L  (most recent actuals; also used as the output template)
  - 2026-27 salary matrix (payroll already worked out)

Output:
  - FY27_Projected_Budget.xlsx  (same layout as the P&L; subtotals stay live
    formulas and recompute when you open the file in Excel / Sheets)
  - a printed summary table

Only dependency: openpyxl  ->  pip install openpyxl
"""

import os
import re
import argparse
from openpyxl import load_workbook

# ============================== CONFIG ==============================
FY_PRIOR1 = "FY25.xlsx"            # FY2024-25  (older)
FY_PRIOR2 = "FY26.xlsx"            # FY2025-26  (most recent actuals + template)
SALARY_MATRIX = "2027_salary_matrix.xlsx"
OUTPUT = "FY27_Projected_Budget.xlsx"

# Folder the spreadsheets live in, relative to the repo root (leave "" if they
# sit at the repo root next to the script). e.g. DATA_DIR = "data"
DATA_DIR = ""

# How NON-salary lines are projected:
#   "yoy_growth" -> apply the FY25->FY26 growth rate to the FY26 actual (capped)
#   "fy26_base"  -> use the FY26 actual as-is
#   "average"    -> average of FY25 and FY26 actuals
#   "flat"       -> FY26 actual * (1 + FLAT_BUMP)
METHOD = "yoy_growth"
GROWTH_CAP = 0.25                  # clamp extrapolated YoY growth to +/- this
FLAT_BUMP = 0.03                   # used only when METHOD == "flat"

# Salaries come from the matrix, not from projection.
SALARY_ACCT = 50610                # 50610 Gross Salaries & Wages
USE_SALARY_FIELD = "total_cost"    # "total_cost" (incl. half bonus) or "new_annual"

# Expense lines that scale with gross wages (multiplied by the salary growth factor).
PAYROLL_SCALED = {50620, 50650, 50660}   # payroll tax, workers comp, employer 401K

PL_SHEET = "Budget vs. Actuals"
SECTIONS = {"Income", "Cost of Goods Sold", "Expenses",
            "Other Income", "Other Expenses"}
# ====================================================================

NUM_RE = re.compile(r"^-?\d+(\.\d+)?$")
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def resolve(name):
    """Find an input file across the likely locations a CI job might run from.
    Raises a FileNotFoundError that lists where it looked and what was there."""
    if os.path.isabs(name):
        if os.path.isfile(name):
            return name
        candidates = [name]
    else:
        bases = [DATA_DIR, os.getcwd(), SCRIPT_DIR,
                 os.path.join(os.getcwd(), "data"),
                 os.path.join(SCRIPT_DIR, "data")]
        seen, candidates = set(), []
        for b in bases:
            if not b:
                continue
            c = os.path.join(b, name)
            if c not in seen:
                seen.add(c)
                candidates.append(c)
        for c in candidates:
            if os.path.isfile(c):
                return c
    listing = []
    for d in dict.fromkeys([os.getcwd(), SCRIPT_DIR]):
        try:
            files = sorted(os.listdir(d)) or ["<empty>"]
        except OSError:
            files = ["<unreadable>"]
        listing.append(f"  {d}:\n    " + "\n    ".join(files))
    raise FileNotFoundError(
        f"Could not find '{name}'. Searched:\n  " + "\n  ".join(candidates)
        + "\nFiles actually present:\n" + "\n".join(listing)
    )


def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip()


def acct_num(label):
    m = re.match(r"\s*(\d{3,6})\b", str(label or ""))
    return int(m.group(1)) if m else None


def leaf_value(raw):
    """Return a float if the cell is a numeric literal ('=99715.47' or 0.0),
    or None if it is empty / a real formula (subtotal, reference, etc.)."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if s.startswith("="):
        s = s[1:].strip()
    return float(s) if NUM_RE.match(s) else None


def read_actuals(path):
    """{normalized_label: {'acct': int|None, 'actual': float}} for leaf rows only."""
    wb = load_workbook(path, data_only=False)
    ws = wb[PL_SHEET] if PL_SHEET in wb.sheetnames else wb.active
    out = {}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        v = leaf_value(ws.cell(r, 2).value)   # column B = Actual
        if v is None:
            continue
        out[norm(a)] = {"acct": acct_num(a), "actual": v}
    return out


def read_salary(path):
    """Return (new_annual_total, total_salary_cost) from the matrix Totals row."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    for r in range(1, ws.max_row + 1):
        if norm(ws.cell(r, 2).value or "") == "Totals":
            new_annual = float(ws.cell(r, 7).value or 0)        # col G
            total_cost = ws.cell(r, 9).value                    # col I
            if not total_cost:                                  # fall back to G + H
                total_cost = new_annual + float(ws.cell(r, 8).value or 0)
            return new_annual, float(total_cost)
    raise ValueError("Could not find the 'Totals' row in the salary matrix.")


def project(label, acct, p1, p2, salary_value, salary_factor):
    if acct == SALARY_ACCT:
        return round(salary_value, 2)

    base2 = p2.get(label, {}).get("actual")   # FY26 actual
    base1 = p1.get(label, {}).get("actual")   # FY25 actual
    if base2 is None and base1 is None:
        return None
    if base2 is None:
        base2 = base1

    if acct in PAYROLL_SCALED:
        return round(base2 * salary_factor, 2)
    if METHOD == "fy26_base":
        return round(base2, 2)
    if METHOD == "average" and base1 is not None:
        return round((base1 + base2) / 2, 2)
    if METHOD == "flat":
        return round(base2 * (1 + FLAT_BUMP), 2)

    # default: yoy_growth
    if base1:
        g = (base2 - base1) / abs(base1)
        g = max(-GROWTH_CAP, min(GROWTH_CAP, g))
        return round(base2 * (1 + g), 2)
    return round(base2, 2)


def build():
    p1 = read_actuals(FY_PRIOR1)
    p2 = read_actuals(FY_PRIOR2)
    new_annual, total_cost = read_salary(SALARY_MATRIX)
    salary_value = total_cost if USE_SALARY_FIELD == "total_cost" else new_annual

    sal_fy26 = next((d["actual"] for d in p2.values()
                     if d["acct"] == SALARY_ACCT), None)
    salary_factor = (salary_value / sal_fy26) if sal_fy26 else 1.0

    # Use FY26 file as the template so structure + subtotal formulas are preserved.
    wb = load_workbook(FY_PRIOR2, data_only=False)
    ws = wb[PL_SHEET] if PL_SHEET in wb.sheetnames else wb.active

    sect = {s: 0.0 for s in SECTIONS}
    current = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        label = norm(a)
        if label in SECTIONS:
            current = label
        if leaf_value(ws.cell(r, 2).value) is None:
            continue   # header / subtotal / total -> leave formula intact
        pv = project(label, acct_num(a), p1, p2, salary_value, salary_factor)
        if pv is None:
            continue
        ws.cell(r, 2).value = pv
        if current:
            sect[current] += pv

    # Relabel headers/title for the projected budget.
    ws.cell(2, 1).value = "Projected Budget: FY2026-27 (generated)"
    ws.cell(3, 1).value = "July 2026 - June 2027"
    if norm(ws.cell(6, 2).value or "").lower() == "actual":
        ws.cell(6, 2).value = "FY27 Projected"
        ws.cell(6, 3).value = "FY26 Budget"
        ws.cell(6, 4).value = "vs FY26 Budget"
        ws.cell(6, 5).value = "% of FY26 Budget"

    wb.save(OUTPUT)

    income = sect["Income"]
    cogs = sect["Cost of Goods Sold"]
    expenses = sect["Expenses"]
    oth_inc = sect["Other Income"]
    oth_exp = sect["Other Expenses"]
    gross = income - cogs
    noi = gross - expenses
    net = noi + oth_inc - oth_exp

    print("FY2026-27 PROJECTED BUDGET  (Marine Exchange of the SF Bay Region)")
    print(f"  Method: {METHOD}   Salary source: matrix '{USE_SALARY_FIELD}' = {salary_value:,.2f}")
    print(f"  Salary scale factor (FY27/FY26 wages): {salary_factor:.4f}")
    print("-" * 52)
    for name, val in [
        ("Total Income", income),
        ("Total Cost of Goods Sold", cogs),
        ("Gross Profit", gross),
        ("Total Expenses", expenses),
        ("Net Operating Income", noi),
        ("Total Other Income", oth_inc),
        ("Total Other Expenses", oth_exp),
        ("NET INCOME", net),
    ]:
        print(f"  {name:<28} {val:>15,.2f}")
    print("-" * 52)
    print(f"  Saved: {OUTPUT}")


def main():
    global FY_PRIOR1, FY_PRIOR2, SALARY_MATRIX, OUTPUT, METHOD
    ap = argparse.ArgumentParser(description="Project FY2026-27 budget.")
    ap.add_argument("--prior1", help="older P&L (FY2024-25)")
    ap.add_argument("--prior2", help="recent P&L (FY2025-26)")
    ap.add_argument("--salary", help="salary matrix xlsx")
    ap.add_argument("--output", help="output xlsx")
    ap.add_argument("--method",
                    choices=["yoy_growth", "fy26_base", "average", "flat"])
    args = ap.parse_args()
    FY_PRIOR1 = args.prior1 or FY_PRIOR1
    FY_PRIOR2 = args.prior2 or FY_PRIOR2
    SALARY_MATRIX = args.salary or SALARY_MATRIX
    OUTPUT = args.output or OUTPUT
    METHOD = args.method or METHOD
    build()


if __name__ == "__main__":
    main()
